import streamlit as st
import pdfplumber
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import re
import json
import os

# ==========================================
# 1. KONFIGURASI HALAMAN & INJEKSI CSS MODERN
# ==========================================
st.set_page_config(page_title="RKAS BOSP Pro", page_icon="🚀", layout="wide")

# Membuat folder fisik permanen jika belum ada
UPLOAD_DIR = "uploads"
if not os.path.exists(UPLOAD_DIR):
    os.makedirs(UPLOAD_DIR)

if 'processed_excel' not in st.session_state:
    st.session_state.processed_excel = None
    st.session_state.processed_filename = ""

# --- SUNTIKAN CSS WEB MODERN (UI/UX UPGRADE) ---
st.markdown("""
    <style>
    /* Mengubah font seluruh web */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;800&display=swap');
    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif;
    }
    
    /* Judul Utama dengan Gradasi Warna Keren */
    .main-title { 
        font-size: 3.5rem; 
        font-weight: 900; 
        background: -webkit-linear-gradient(45deg, #1E88E5, #D81B60);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        text-align: center;
        margin-bottom: 0.5rem;
        padding-top: 1rem;
    }
    
    /* Sub-judul yang elegan */
    .sub-title { 
        font-size: 1.2rem; 
        color: #6c757d; 
        text-align: center;
        margin-bottom: 3rem;
        font-weight: 400;
    }
    
    /* Styling Area Upload (Dropzone) agar modern */
    [data-testid="stFileUploadDropzone"] {
        background-color: #f8f9fa;
        border: 2px dashed #1E88E5;
        border-radius: 15px;
        padding: 2rem;
        transition: all 0.3s cubic-bezier(0.25, 0.8, 0.25, 1);
    }
    [data-testid="stFileUploadDropzone"]:hover {
        background-color: #e3f2fd;
        border-color: #D81B60;
        transform: translateY(-2px);
        box-shadow: 0 8px 20px rgba(30, 136, 229, 0.15);
    }
    
    /* Animasi Tombol Download / Primary */
    .stButton > button {
        border-radius: 8px !important;
        font-weight: 600 !important;
        transition: all 0.3s ease !important;
    }
    .stButton > button:hover {
        transform: translateY(-3px);
        box-shadow: 0 10px 20px rgba(0,0,0,0.1) !important;
    }
    
    /* Merapikan Tab */
    .stTabs [data-baseweb="tab-list"] { 
        gap: 2rem; 
        justify-content: center;
        margin-bottom: 2rem;
    }
    .stTabs [data-baseweb="tab"] { 
        height: 3.5rem; 
        font-size: 1.1rem; 
        font-weight: 600; 
        border-radius: 8px 8px 0 0;
    }
    
    /* Sidebar Styling */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #f4f6f9 0%, #ffffff 100%);
        border-right: 1px solid #e0e0e0;
    }
    </style>
""", unsafe_allow_html=True)

# ==========================================
# 2. SISTEM MANAJEMEN LINK (PORTAL ADMIN)
# ==========================================
LINKS_FILE = "links.json"

def load_links():
    if os.path.exists(LINKS_FILE):
        try:
            with open(LINKS_FILE, "r") as f:
                return json.load(f)
        except: pass
    return {"Aplikasi Utama RKAS Manual": "https://script.google.com/macros/s/AKfycbzxWVN-UBAf-VimxCFg4CPgWhRu6pqpy0qlzSrWJEja44mvnVZE1c680fQnuohQLYMo/exec"}

def save_links(links):
    with open(LINKS_FILE, "w") as f:
        json.dump(links, f)

if 'app_links' not in st.session_state:
    st.session_state.app_links = load_links()

# ==========================================
# 3. SIDEBAR (PANEL KIRI)
# ==========================================
with st.sidebar:
    # Memakai kolom agar gambar rapi di tengah
    col_img1, col_img2, col_img3 = st.columns([1,2,1])
    with col_img2:
        st.image("https://cdn-icons-png.flaticon.com/512/3135/3135692.png", use_container_width=True) 
    
    st.markdown("<h2 style='text-align: center; color: #1E88E5;'>Navigasi Sistem</h2>", unsafe_allow_html=True)
    
    # --- 1. TOMBOL PANDUAN (POP-UP) ---
    with st.popover("📖 BACA PANDUAN SISTEM", use_container_width=True):
        st.markdown("""
        **Pilih jenis dokumen yang ingin Anda rapikan di halaman utama:**
        
        📄 **MODE PDF:** Ubah PDF Kertas Kerja BOSP langsung menjadi format Excel rapi.
        
        📊 **MODE CSV / EXCEL KOTOR:** Rapikan file Excel/CSV hasil *convert* dari aplikasi lain.
        """)
        st.info("⚡ **Info:** Total menggunakan sistem Hybrid yang dijamin presisi dan anti-error.")

    # --- 2. RUANG TITIP FILE PERMANEN ---
    st.divider()
    st.markdown("### 🗂️ Ruang Titip (Permanen)")
    st.caption("Upload file pendukung di sini. File akan tersimpan selamanya.")
    
    titipan_file = st.file_uploader("Upload File Bebas", key="free_uploader", label_visibility="collapsed")
    
    if titipan_file is not None:
        file_path = os.path.join(UPLOAD_DIR, titipan_file.name)
        with open(file_path, "wb") as f:
            f.write(titipan_file.getbuffer())
        st.success(f"✅ '{titipan_file.name}' tersimpan!")
        
    saved_files = os.listdir(UPLOAD_DIR)
    if saved_files:
        st.markdown("**📂 File Tersedia:**")
        for file_name in saved_files:
            file_path = os.path.join(UPLOAD_DIR, file_name)
            with open(file_path, "rb") as f:
                file_bytes = f.read()
            st.download_button(
                label=f"⬇️ {file_name}",
                data=file_bytes,
                file_name=file_name,
                use_container_width=True,
                key=f"dl_perm_{file_name}"
            )
    else:
        st.info("Belum ada file yang dititipkan.")
    
    # --- TEMPAT NAVIGASI UNDUHAN HASIL EXCEL ---
    download_section = st.empty()
    
    # --- MENAMPILKAN SEMUA LINK (PORTAL APLIKASI) ---
    st.divider()
    st.markdown("### 🌐 Portal Aplikasi")
    for nama_link, url_link in st.session_state.app_links.items():
        st.link_button(f"🚀 {nama_link}", url_link, use_container_width=True)
        
    st.divider()
    
    # --- PORTAL ADMIN (KELOLA LINK & FILE) ---
    with st.expander("⚙️ Akses Admin (Link & File)"):
        admin_pwd = st.text_input("Masukkan Password", type="password")
        if admin_pwd == "admin123":
            st.success("Akses Terbuka!")
            
            st.markdown("**1. Tambah Link Baru**")
            new_name = st.text_input("Nama Tombol")
            new_url = st.text_input("Alamat URL (https://...)")
            
            if st.button("➕ Simpan Link", use_container_width=True):
                if new_name and new_url:
                    st.session_state.app_links[new_name] = new_url
                    save_links(st.session_state.app_links)
                    st.success("Tersimpan!")
                    st.rerun()
                else: st.error("Tidak boleh kosong.")
            
            st.markdown("---")
            st.markdown("**2. Hapus Link Lama**")
            link_to_delete = st.selectbox("Pilih link:", options=[""] + list(st.session_state.app_links.keys()))
            if st.button("🗑️ Hapus Link", use_container_width=True):
                if link_to_delete and link_to_delete in st.session_state.app_links:
                    del st.session_state.app_links[link_to_delete]
                    save_links(st.session_state.app_links)
                    st.success("Dihapus!")
                    st.rerun()
                    
            st.markdown("---")
            st.markdown("**3. Bersihkan File Server**")
            file_to_delete = st.selectbox("Pilih file:", options=[""] + os.listdir(UPLOAD_DIR))
            if st.button("🧨 Hapus File Fisik", use_container_width=True):
                if file_to_delete:
                    path_to_delete = os.path.join(UPLOAD_DIR, file_to_delete)
                    if os.path.exists(path_to_delete):
                        os.remove(path_to_delete)
                        st.success("Musnah!")
                        st.rerun()

# ==========================================
# 4. FUNGSI MERAPIKAN EXCEL (HYBRID CALCULATION)
# ==========================================
def create_styled_excel(cleaned_data):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rincian Kertas Kerja"
    
    ws.append(["No. Urut", "Kode Rekening", "Kode Program", "Uraian", "Rincian Perhitungan", "", "", "Jumlah"])
    ws.append(["", "", "", "", "Volume", "Satuan", "Tarif Harga", ""])
    
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:C2')
    ws.merge_cells('D1:D2')
    ws.merge_cells('E1:G1')
    ws.merge_cells('H1:H2')
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid") 
    
    for r_idx in range(1, 3):
        for c_idx in range(1, 9):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
    
    for row_idx, row_data in enumerate(cleaned_data, start=3):
        is_grand_total = "jumlah" in str(row_data[3]).lower() or "jumlah" in str(row_data[0]).lower()
        is_item = bool(str(row_data[1]).strip()) 
        
        for col_idx, val_str in enumerate(row_data, start=1):
            val_bersih = str(val_str).replace('\n', ' ').strip()
            
            if not val_bersih: cell = ws.cell(row=row_idx, column=col_idx, value=None)
            else: cell = ws.cell(row=row_idx, column=col_idx, value=val_bersih)
                
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            if val_bersih and col_idx in [5, 7, 8]:
                try:
                    num_str = val_bersih.replace('Rp', '').replace(' ', '')
                    num_str = num_str.replace('.', '').replace(',', '.')
                    num_val = float(num_str)
                    
                    if col_idx == 5: cell.value = num_val
                    elif col_idx == 7: 
                        cell.value = num_val
                        cell.number_format = '#,##0'
                    elif col_idx == 8 and not is_grand_total and not is_item:
                        cell.value = num_val
                        cell.number_format = '#,##0'
                except: pass
            
            if col_idx == 8:
                if is_grand_total:
                    cell.value = f'=SUMIF(B3:B{row_idx-1}, "<>", H3:H{row_idx-1})'
                    cell.number_format = '#,##0'
                elif is_item:
                    try:
                        val_e = ws.cell(row=row_idx, column=5).value
                        val_g = ws.cell(row=row_idx, column=7).value
                        if isinstance(val_e, (int, float)) and isinstance(val_g, (int, float)):
                            cell.value = val_e * val_g
                        else:
                            n_str = val_bersih.replace('Rp', '').replace(' ', '').replace('.', '').replace(',', '.')
                            cell.value = float(n_str)
                    except:
                        try:
                            n_str = val_bersih.replace('Rp', '').replace(' ', '').replace('.', '').replace(',', '.')
                            cell.value = float(n_str)
                        except: cell.value = val_bersih
                    cell.number_format = '#,##0'
                
            if not is_item and not is_grand_total:
                cell.font = bold_font
                
            if col_idx in [1, 2, 3, 5, 6]:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 16
    
    for r_idx in range(3, ws.max_row + 1):
        cell_d = ws.cell(row=r_idx, column=4)
        if cell_d.value and "jumlah" in str(cell_d.value).lower():
            cell_d.font = bold_font
            cell_d.alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row=r_idx, column=8).font = bold_font
            
    wb.save(output)
    return output.getvalue()


# ==========================================
# 5. KONTEN UTAMA (TABS & UI MODERN)
# ==========================================
st.markdown('<p class="main-title">Aplikasi Konverter BOSP ⚡</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-title">Mesin Pintar Pengolah Kertas Kerja RKAS Menjadi Excel Sempurna</p>', unsafe_allow_html=True)

tab1, tab2 = st.tabs(["📄 MODE PDF (Auto-Scan)", "📊 MODE CSV (Data Recovery)"])

# ------------------------------------------
# TAB 1: MODE PDF LAMA
# ------------------------------------------
with tab1:
    st.markdown("<h4 style='text-align: center; color: #424242; margin-bottom: 20px;'>Seret & Lepas File Kertas Kerja (PDF) Anda Di Sini</h4>", unsafe_allow_html=True)
    uploaded_pdf = st.file_uploader("", type="pdf", key="pdf_uploader", label_visibility="collapsed")
    
    if uploaded_pdf is not None:
        with st.spinner("⏳ Membedah dan Menganalisis Dokumen..."):
            try:
                raw_rows = []
                with pdfplumber.open(uploaded_pdf) as pdf:
                    for page in pdf.pages:
                        table = page.extract_table()
                        if table: raw_rows.extend(table)
                
                if raw_rows:
                    cleaned_data = []
                    for row in raw_rows:
                        if not row: continue
                        cells = [str(x).strip() for x in row if x is not None and str(x).strip() != ""]
                        if not cells: continue
                        
                        row_str_lower = " ".join(cells).lower()
                        
                        is_garbage = False
                        if any(kw in row_str_lower for kw in ["npsn", "nama sekolah", "alamat", "kabupaten", "provinsi", "sumber dana", "total penerimaan"]): 
                            is_garbage = True
                        if "kode rekening" in row_str_lower and "uraian" in row_str_lower: is_garbage = True
                        if "rincian perhitungan" in row_str_lower and "tarif harga" in row_str_lower: is_garbage = True
                        if str(cells[0]).lower().strip() in ["a. penerimaan", "b. belanja"]: is_garbage = True
                        if "volume" in row_str_lower and "tarif harga" in row_str_lower and "satuan" in row_str_lower: is_garbage = True
                        
                        if is_garbage: continue
                        if len(cells) == 1 and "jumlah" not in row_str_lower: continue
                        
                        no_urut, kode_rek, kode_prog, uraian, volume, satuan, tarif, jumlah = "", "", "", "", "", "", "", ""
                        
                        if "jumlah" in row_str_lower and len(cells) <= 3:
                            uraian = "Jumlah"
                            for c in reversed(cells):
                                if sum(char.isdigit() for char in c) > 3:
                                    jumlah = c
                                    break
                            cleaned_data.append([no_urut, kode_rek, kode_prog, uraian, volume, satuan, tarif, jumlah])
                            continue
                        
                        if re.match(r'^\d+\.$', cells[0]): no_urut = cells.pop(0)
                        if cells and re.match(r'^5\.\d', cells[0]): kode_rek = cells.pop(0)
                        
                        prog_parts = []
                        while cells and re.match(r'^(?:\d{2}\.\s*)+$', cells[0]):
                            prog_parts.append(cells.pop(0))
                        if prog_parts: kode_prog = " ".join(prog_parts)
                        
                        if cells and re.match(r'^((?:\d{2}\.\s*)+)\s*(.+)$', cells[0]):
                            m = re.match(r'^((?:\d{2}\.\s*)+)\s*(.+)$', cells[0])
                            kode_prog += (" " if kode_prog else "") + m.group(1).strip()
                            if m.group(2).strip(): cells[0] = m.group(2).strip()
                            else: cells.pop(0)
                        
                        if not cells: continue
                            
                        if kode_rek:
                            if len(cells) >= 5:
                                uraian = " ".join(cells[:-4])
                                volume, satuan, tarif, jumlah = cells[-4], cells[-3], cells[-2], cells[-1]
                            elif len(cells) == 4:
                                m = re.match(r'^([\d\.,]+)\s+([a-zA-Z\s/]+)$', cells[-3])
                                if m:
                                    uraian = " ".join(cells[:-3])
                                    volume, satuan = m.group(1).strip(), m.group(2).strip()
                                    tarif, jumlah = cells[-2], cells[-1]
                                else:
                                    uraian = " ".join(cells[:-2])
                                    tarif, jumlah = cells[-2], cells[-1]
                            else: uraian = " ".join(cells)
                        else:
                            if len(cells) >= 2:
                                uraian = " ".join(cells[:-1])
                                jumlah = cells[-1]
                            elif len(cells) == 1: uraian = cells[0]
                        
                        if any([no_urut, kode_rek, kode_prog, uraian, volume, satuan, tarif, jumlah]):
                            cleaned_data.append([no_urut, kode_rek, kode_prog, uraian, volume, satuan, tarif, jumlah])
                    
                    if cleaned_data:
                        st.success(f"✅ Berhasil mengekstrak {len(cleaned_data)} baris data!")
                        df_preview = pd.DataFrame(cleaned_data, columns=["No. Urut", "Kode Rekening", "Kode Program", "Uraian", "Volume", "Satuan", "Tarif Harga", "Jumlah"])
                        st.dataframe(df_preview, use_container_width=True, height=350)
                        
                        st.session_state.processed_excel = create_styled_excel(cleaned_data)
                        st.session_state.processed_filename = "RKAS_BOSP_PDF_Rapi.xlsx"
                        
                        st.info("✨ Proses Selesai! File Excel Anda sudah siap diunggah. Silakan klik tombol Download di panel sebelah kiri.")
                    else: st.warning("Data tabel kosong.")
                else: st.warning("Gagal membaca isi tabel pada PDF.")
            except Exception as e: st.error(f"Error pada sistem: {e}")

# ------------------------------------------
# TAB 2: MODE CSV / EXCEL KOTOR
# ------------------------------------------
with tab2:
    st.markdown("<h4 style='text-align: center; color: #424242; margin-bottom: 20px;'>Unggah File CSV / Excel (XLSX) Kotor Anda Di Sini</h4>", unsafe_allow_html=True)
    uploaded_csv = st.file_uploader("", type=["csv", "xlsx", "xls"], key="csv_uploader", label_visibility="collapsed")
    
    if uploaded_csv is not None:
        with st.spinner("⏳ Memperbaiki Format dan Menghancurkan Karakter Rusak..."):
            try:
                if uploaded_csv.name.endswith('.csv'): df_raw = pd.read_csv(uploaded_csv, header=None)
                else: df_raw = pd.read_excel(uploaded_csv, header=None)
                
                df_raw = df_raw.fillna("").astype(str)
                cleaned_csv_data = []
                table_started = False 
                
                for index, row in df_raw.iterrows():
                    r = [str(x).replace('nan', '').replace('NaN', '').strip() for x in row.tolist()]
                    while len(r) < 8: r.append("")
                    
                    row_str_lower = " ".join(r).lower()
                    if not table_started:
                        if "kode rekening" in row_str_lower or "b. belanja" in row_str_lower: table_started = True
                        continue
                    
                    if "kode rekening" in str(r[1]).lower() or "kode rekening" in str(r[0]).lower(): continue
                    if "uraian" in str(r[3]).lower() or "rincian perhitungan" in str(r[3]).lower(): continue
                    if "no. urut" in str(r[0]).lower() or "b. belanja" in str(r[0]).lower(): continue
                    if "a. penerimaan" in str(r[0]).lower(): continue
                    if "volume" in row_str_lower and "tarif harga" in row_str_lower and "satuan" in row_str_lower: continue
                    
                    non_empty_r = [x for x in r if x]
                    if len(non_empty_r) == 1 and "jumlah" not in row_str_lower: continue
                    if len(non_empty_r) == 0: continue
                    
                    no_urut, kode_rek, kode_prog, uraian, volume, satuan, tarif, jumlah = "", "", "", "", "", "", "", ""
                    
                    if "jumlah" in row_str_lower and len(non_empty_r) <= 3:
                        uraian = "Jumlah"
                        for c in reversed(r):
                            if c and sum(char.isdigit() for char in c) > 3:
                                jumlah = c
                                break
                        cleaned_csv_data.append([no_urut, kode_rek, kode_prog, uraian, volume, satuan, tarif, jumlah])
                        continue
                    
                    no_urut, kode_rek = r[0], r[1]
                    kode_prog_mentah = r[2]
                    
                    m_date = re.match(r'^(\d{4})-(\d{2})-(\d{2})(?:\s.*)?$', kode_prog_mentah)
                    if m_date:
                        yyyy, mm, dd = m_date.groups()
                        if int(yyyy) >= 2024: kode_prog = f"{dd}. {mm}."
                        else: kode_prog = f"{dd}. {mm}. {yyyy[-2:]}."
                    elif re.match(r'^(\d{2})[/-](\d{2})[/-](\d{4})(?:\s.*)?$', kode_prog_mentah):
                        m_date2 = re.match(r'^(\d{2})[/-](\d{2})[/-](\d{4})(?:\s.*)?$', kode_prog_mentah)
                        p1, p2, yyyy = m_date2.groups()
                        if int(yyyy) >= 2024: kode_prog = f"{p1}. {p2}."
                        else: kode_prog = f"{p1}. {p2}. {yyyy[-2:]}."
                    else: kode_prog = re.sub(r'\s+00:00:00$', '', kode_prog_mentah)
                    
                    uraian = r[3]
                    rem = [x for x in r[4:] if x]
                    if not uraian and rem: uraian = rem.pop(0)
                        
                    if kode_rek: 
                        if len(rem) >= 4:
                            if len(rem) > 4: uraian += " " + " ".join(rem[:-4])
                            volume, satuan, tarif, jumlah = rem[-4], rem[-3], rem[-2], rem[-1]
                        elif len(rem) == 3:
                            m = re.match(r'^([\d\.,]+)\s*(.*)$', rem[0])
                            if m: volume, satuan = m.group(1).strip(), m.group(2).strip()
                            else: volume = rem[0]
                            tarif, jumlah = rem[1], rem[2]
                        elif len(rem) == 2: tarif, jumlah = rem[0], rem[1]
                    else: 
                        if len(rem) >= 1:
                            jumlah = rem[-1]
                            if len(rem) > 1: uraian += " " + " ".join(rem[:-1])
                                
                    if any([no_urut, kode_rek, kode_prog, uraian, volume, satuan, tarif, jumlah]):
                        cleaned_csv_data.append([no_urut, kode_rek, kode_prog, uraian, volume, satuan, tarif, jumlah])

                if cleaned_csv_data:
                    st.success(f"✅ Data Terselamatkan! Total: {len(cleaned_csv_data)} baris.")
                    df_preview_csv = pd.DataFrame(cleaned_csv_data, columns=["No. Urut", "Kode Rekening", "Kode Program", "Uraian", "Volume", "Satuan", "Tarif Harga", "Jumlah"])
                    st.dataframe(df_preview_csv, use_container_width=True, height=350)
                    
                    st.session_state.processed_excel = create_styled_excel(cleaned_csv_data)
                    st.session_state.processed_filename = "RKAS_BOSP_Dari_CSV_Rapi.xlsx"
                    
                    st.info("✨ Proses Selesai! File Excel Anda sudah siap diunggah. Silakan klik tombol Download di panel sebelah kiri.")
                else: st.warning("Gagal menyaring data. Pastikan format dokumennya mirip Kertas Kerja.")
            except Exception as e: st.error(f"Terjadi kesalahan saat memproses CSV/Excel: {e}")

# ==========================================
# 6. MENGISI PLACEHOLDER PUSAT UNDUHAN DI SIDEBAR
# ==========================================
with download_section.container():
    if st.session_state.get('processed_excel'):
        st.markdown(
            """
            <div style="background: -webkit-linear-gradient(45deg, #43a047, #1b5e20); padding: 15px; border-radius: 10px; color: white; text-align: center; margin-bottom: 10px; font-weight: bold; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
            🎉 EXCEL SUDAH JADI!
            </div>
            """, unsafe_allow_html=True
        )
        st.download_button(
            label="⬇️ DOWNLOAD EXCEL FINAL",
            data=st.session_state.processed_excel,
            file_name=st.session_state.processed_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
